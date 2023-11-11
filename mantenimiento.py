import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook

class MantenimientoVehiculosApp:
    def _init_(self, master):
        self.master = master
        self.master.tlitle("mantenimiento de vehiculos")

        #crear etiquetas de entradas
        self.label_codigo = tk.label (master, text="codigo:")
        self.label_marca = tk.label (master, text="marca:")
        self.label_modelo = tk.label (master, text="modelo:")
        self.label_precio = tk.label (master, text="precio:")
        self.label_kilometraje = tk.label (master, text="kilometraje:")

        self.Entry_codigo = tk.Entry(master)
        self.Entry_marca = tk.Entry(master)
        self.Entry_modelo = tk.Entry(master)
        self.Entry_precio = tk.Entry(master)
        self.Entry_kilometraje = tk.Entry(master)

        #botones
        self.btn_guardar = tk.button(master, text="Guardar", command=self.guardar_vehiculos) 
        self.btn_editar = tk.button(master, text="editar", command=self.editar_vehiculos) 
        self.btn_eliminar = tk.button(master, text="eliminar", command=self.eliminar_vehiculos) 
        self.btn_listar = tk.button(master, text="listar", command=self.listar_vehiculos) 

        #posicionamiento de elementos en la ventana
        self.label_codigo.grid(row=0, column=0)
        self.label_marca.grid(row=1, column=0)
        self.label_modelo.grid(row=2, column=0)
        self.label_precio.grid(row=3, column=0)
        self.label_kilometraje.grid(row=4, column=0)

        self.entry_codigo.grid(row=0, columna=1)
        self.entry_marca.grid(row=1, columna=1)
        self.entry_modelo.grid(row=2, columna=1)
        self.entry_precio.grid(row=3, columna=1)
        self.entry_kilometraje.grid(row=4, columna=1)

        self.btn_guardar.grid(row=5, column=0, columnsan=2, pady=10)
        self.btn_editar.grid(row=6, column=0, columnsan=2, pady=10)
        self.btn_eliminar.grid(row=7, column=0, columnsan=2, pady=10)
        self.btn_listar.grid(row=8, column=0, columnsan=2, pady=10)

        def guardar_vehiculos(self):
            codigo = self.entry_codigo.get()
            marca = self.entry_marca.get()
            modelo = self.entry_modelo.get()
            precio = self.entry_precio.get()
            kilometraje = self.entry_kilometraje.get()

            if codigo and marca and modelo and precio and kilometraje:
                #abrir el archivo en Excel
                try:
                    wb = load_workbook("vehiculos.xlsx")
                    sheet = wb ["listado"]
                except FileNotFoundError:
                    wb = load_workbook("vehiculos.xlsx")
                    sheet = wb.create_sheet("listado")

                    #escribir datos en el archivo Excel
                    sheet.append([codigo, marca, modelo, float(precio), int (kilometraje)])
                    wb.sabe("vechiculos.xlsx")
                    


